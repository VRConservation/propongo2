"""Main Flask application for Propongo2."""

import json
import os
import re
import math
import uuid
import logging
import threading
from datetime import datetime
from typing import Any, Optional, Tuple

from flask import Flask, render_template, request, jsonify, redirect, url_for, Response
from markupsafe import Markup
import markdown
from .models import Proposal, PROPOSALS_DIR
from .export import export_bp
from .snippets import snippets_bp
from .utils import build_export_context
from .config import Config, ERROR_MESSAGES
from . import __version__

# Configure logging
logging.basicConfig(
    level=getattr(logging, Config.LOG_LEVEL),
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

_proposal_locks = {}
_proposal_locks_lock = threading.Lock()


def markdown_to_html(text: str) -> str:
    """Convert Markdown text to HTML using the markdown library.
    
    Args:
        text: Markdown formatted text
        
    Returns:
        HTML formatted string
    """
    if not text:
        return ""
    return markdown.markdown(
        text,
        extensions=['tables', 'nl2br', 'fenced_code', 'sane_lists']
    )


def validate_numeric(value: Any, name: str, min_val: float = 0.0) -> float:
    """Validate and convert a numeric value.
    
    Args:
        value: Value to validate
        name: Name of the field (for error messages)
        min_val: Minimum allowed value
        
    Returns:
        Validated float value
        
    Raises:
        ValueError: If value is invalid
    """
    try:
        num = float(value)
        if not math.isfinite(num):
            raise ValueError(f"{name} must be a finite number")
        if num < min_val:
            raise ValueError(f"{name} must be >= {min_val}")
        return num
    except (ValueError, TypeError) as e:
        raise ValueError(f"Invalid {name}: {str(e)}")


def create_app() -> Flask:
    """Create and configure the Flask application.
    
    Returns:
        Configured Flask application
    """
    app = Flask(__name__)
    app.secret_key = Config.SECRET_KEY
    
    logger.info(f"Starting Propongo2 v{__version__}")

    app.register_blueprint(export_bp)
    app.register_blueprint(snippets_bp)

    @app.context_processor
    def inject_version():
        return {"app_version": __version__}

    app.jinja_env.filters["md"] = lambda text: Markup(markdown_to_html(text))
    app.jinja_env.filters["currency"] = lambda value: f"{value:,.0f}"

    @app.route("/")
    def index():
        proposals = Proposal.list_all()
        return render_template("index.html", proposals=proposals)

    @app.route("/new")
    def new_proposal():
        proposal = Proposal()
        proposal.save()
        logger.info(f"Created new proposal: {proposal.id}")
        return redirect(url_for("editor", proposal_id=proposal.id))

    @app.route("/editor/<proposal_id>")
    def editor(proposal_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return redirect(url_for("index"))
        return render_template(
            "base.html",
            proposal=proposal,
            tasks=proposal.tasks,
            budget_items=proposal.budget_items,
        )

    @app.route("/api/proposal/<proposal_id>", methods=["GET"])
    def get_proposal(proposal_id: str) -> Tuple[Response, int]:
        proposal = Proposal.load(proposal_id)
        if not proposal:
            logger.warning(f"Proposal not found: {proposal_id}")
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404
        return jsonify(proposal.to_dict()), 200

    @app.route("/api/proposal/<proposal_id>", methods=["PUT"])
    def save_proposal(proposal_id: str) -> Tuple[Response, int]:
        data = request.get_json()
        if not data:
            return jsonify(ERROR_MESSAGES['NO_DATA']), 400

        with _proposal_locks_lock:
            if proposal_id not in _proposal_locks:
                _proposal_locks[proposal_id] = threading.Lock()
            lock = _proposal_locks[proposal_id]

        with lock:
            proposal = Proposal.load(proposal_id)
            if not proposal:
                logger.warning(f"Proposal not found for update: {proposal_id}")
                return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404

            if "tasks" in data:
                incoming_tasks = data.pop("tasks")
                existing_by_id = {t.get("id"): t for t in proposal.tasks}
                merged = []
                for t in incoming_tasks:
                    tid = t.get("id", "")
                    if tid in existing_by_id:
                        merged_task = dict(existing_by_id[tid])
                        merged_task.update(t)
                    else:
                        merged_task = t
                    merged.append(merged_task)
                proposal.tasks = merged

            skip_fields = {"id", "title", "created_at"}
            for key, value in data.items():
                if key in skip_fields:
                    continue
                if hasattr(proposal, key):
                    setattr(proposal, key, value)

            proposal.save()
            return jsonify(proposal.to_dict())

    @app.route("/api/proposal/<proposal_id>", methods=["DELETE"])
    def delete_proposal(proposal_id):
        Proposal.delete(proposal_id)
        return jsonify({"ok": True})

    @app.route("/api/proposal/<proposal_id>/save-as", methods=["POST"])
    def save_proposal_as(proposal_id):
        data = request.get_json()
        title = data.get("title", "").strip()
        if not title:
            return jsonify({"error": "Title required"}), 400

        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify({"error": "Not found"}), 404

        new_id = re.sub(r"[^a-z0-9_-]", "_", title.lower())
        new_id = re.sub(r"_+", "_", new_id).strip("_")
        if not new_id:
            new_id = uuid.uuid4().hex[:8]

        original_id = new_id
        counter = 1
        while Proposal.load(new_id):
            new_id = f"{original_id}_{counter}"
            counter += 1

        new_proposal = Proposal(id=new_id, title=title)
        new_proposal.client_name = proposal.client_name
        new_proposal.subtitle = getattr(proposal, 'subtitle', '') or ''
        new_proposal.project_summary = proposal.project_summary
        new_proposal.scope = getattr(proposal, 'scope', '') or ''
        new_proposal.tasks = list(proposal.tasks)
        new_proposal.qualifications = proposal.qualifications
        new_proposal.budget_items = list(proposal.budget_items)
        new_proposal.budget_item_timings = dict(proposal.budget_item_timings) if proposal.budget_item_timings else {}
        new_proposal.start_date = proposal.start_date
        new_proposal.indirect_percent = getattr(proposal, 'indirect_percent', 0) or 0
        new_proposal.show_budget_description = getattr(proposal, 'show_budget_description', False)
        new_proposal.budget_description = getattr(proposal, 'budget_description', '') or ''
        new_proposal.custom_sections = list(proposal.custom_sections) if proposal.custom_sections else []
        new_proposal.timeline_use_days = proposal.timeline_use_days
        new_proposal.timeline_show_budget = proposal.timeline_show_budget
        new_proposal.end_date = getattr(proposal, 'end_date', '') or ''
        new_proposal.milestones = list(proposal.milestones) if proposal.milestones else []
        new_proposal.reports = list(proposal.reports) if proposal.reports else []
        new_proposal.save()

        return jsonify({"id": new_id}), 201

    @app.route("/api/proposals", methods=["GET"])
    def list_proposals():
        return jsonify(Proposal.list_all())

    @app.route("/templates")
    def templates_page():
        templates = Proposal.list_templates()
        return render_template("templates.html", templates=templates)

    @app.route("/api/templates", methods=["GET"])
    def list_templates():
        return jsonify(Proposal.list_templates())

    @app.route("/api/template/<template_id>", methods=["DELETE"])
    def delete_template(template_id):
        Proposal.delete(template_id, is_template=True)
        return jsonify({"ok": True})

    @app.route("/api/proposal/<proposal_id>/save-as-template", methods=["POST"])
    def save_as_template(proposal_id):
        data = request.get_json()
        template_name = data.get("template_name", "").strip()
        template_category = data.get("template_category", "").strip()
        if not template_name:
            return jsonify({"error": "Template name required"}), 400

        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify({"error": "Not found"}), 404

        new_id = re.sub(r"[^a-z0-9_-]", "_", template_name.lower())
        new_id = re.sub(r"_+", "_", new_id).strip("_")
        if not new_id:
            new_id = uuid.uuid4().hex[:8]

        original_id = new_id
        counter = 1
        while Proposal.load(new_id, is_template=True):
            new_id = f"{original_id}_{counter}"
            counter += 1

        tmpl = Proposal(id=new_id, title=proposal.title)
        tmpl.client_name = proposal.client_name
        tmpl.subtitle = getattr(proposal, 'subtitle', '') or ''
        tmpl.project_summary = proposal.project_summary
        tmpl.scope = getattr(proposal, 'scope', '') or ''
        tmpl.tasks = list(proposal.tasks)
        tmpl.qualifications = proposal.qualifications
        tmpl.budget_items = list(proposal.budget_items)
        tmpl.budget_item_timings = dict(proposal.budget_item_timings) if proposal.budget_item_timings else {}
        tmpl.indirect_percent = getattr(proposal, 'indirect_percent', 0) or 0
        tmpl.show_budget_description = getattr(proposal, 'show_budget_description', False)
        tmpl.budget_description = getattr(proposal, 'budget_description', '') or ''
        tmpl.custom_sections = list(proposal.custom_sections) if proposal.custom_sections else []
        tmpl.timeline_use_days = proposal.timeline_use_days
        tmpl.timeline_show_budget = proposal.timeline_show_budget
        tmpl.is_template = True
        tmpl.template_name = template_name
        tmpl.template_category = template_category
        tmpl.save()

        return jsonify({"id": new_id}), 201

    @app.route("/templates/new-from/<template_id>")
    def new_from_template(template_id):
        tmpl = Proposal.load(template_id, is_template=True)
        if not tmpl:
            return redirect(url_for("templates_page"))

        new_id = uuid.uuid4().hex[:8]
        proposal = Proposal(id=new_id, title=tmpl.title)
        proposal.client_name = tmpl.client_name
        proposal.subtitle = getattr(tmpl, 'subtitle', '') or ''
        proposal.project_summary = tmpl.project_summary
        proposal.scope = getattr(tmpl, 'scope', '') or ''
        proposal.tasks = list(tmpl.tasks)
        proposal.qualifications = tmpl.qualifications
        proposal.budget_items = list(tmpl.budget_items)
        proposal.budget_item_timings = dict(tmpl.budget_item_timings) if tmpl.budget_item_timings else {}
        proposal.indirect_percent = getattr(tmpl, 'indirect_percent', 0) or 0
        proposal.show_budget_description = getattr(tmpl, 'show_budget_description', False)
        proposal.budget_description = getattr(tmpl, 'budget_description', '') or ''
        proposal.custom_sections = list(tmpl.custom_sections) if tmpl.custom_sections else []
        proposal.timeline_use_days = tmpl.timeline_use_days
        proposal.timeline_show_budget = tmpl.timeline_show_budget
        proposal.save()

        return redirect(url_for("editor", proposal_id=proposal.id))

    @app.route("/scope/<proposal_id>")
    def scope_tab(proposal_id: str) -> Tuple[str, int] | Tuple[Response, int]:
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404
        return render_template("scope.html", proposal=proposal, tasks=proposal.tasks)

    @app.route("/budget/<proposal_id>")
    def budget_tab(proposal_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404

        task_budgets = {}
        for task in proposal.tasks:
            items = [b for b in proposal.budget_items if b.get("task_id") == task["id"]]
            subtotal = sum(i.get("cost_per_unit", 0) * i.get("units", 0) for i in items)
            task_budgets[task["id"]] = {
                "task": task,
                "items": items,
                "subtotal": subtotal,
            }

        indirect_percent = getattr(proposal, 'indirect_percent', 0) or 0
        indirect_amount = proposal.total_budget * (indirect_percent / 100)
        total_with_indirect = proposal.total_budget + indirect_amount

        return render_template(
            "budget.html",
            proposal=proposal,
            tasks=proposal.tasks,
            budget_items=proposal.budget_items,
            total_budget=proposal.total_budget,
            task_budgets=task_budgets,
            indirect_percent=indirect_percent,
            indirect_amount=indirect_amount,
            total_with_indirect=total_with_indirect,
        )

    @app.route("/qualifications/<proposal_id>")
    def qualifications_tab(proposal_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404
        return render_template("qualifications.html", proposal=proposal)

    @app.route("/timeline/<proposal_id>")
    def timeline_tab(proposal_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404

        task_budgets = {}
        timings = proposal.budget_item_timings or {}
        for task in proposal.tasks:
            items = [b for b in proposal.budget_items if b.get("task_id") == task["id"]]
            for item in items:
                t = timings.get(item.get("id", ""), {})
                if t:
                    item["start_month"] = t.get("start_month")
                    item["start_year"] = t.get("start_year")
                    item["duration_months"] = t.get("duration_months", 1)
                    if t.get("lead_entity"):
                        item["lead_entity"] = t["lead_entity"]
                    item["recurring"] = t.get("recurring", False)
                    item["recurring_interval"] = t.get("recurring_interval", 3)
            if items:
                task_budgets[task["id"]] = {
                    "task": task,
                    "items": items,
                }

        try:
            sd = datetime.strptime(proposal.start_date, "%Y-%m-%d")
            start_date_month = sd.month
            start_date_year = sd.year
        except (ValueError, TypeError):
            now = datetime.now()
            start_date_month = now.month
            start_date_year = now.year

        try:
            ed = datetime.strptime(proposal.end_date, "%Y-%m-%d") if proposal.end_date else None
            end_date_month = ed.month if ed else start_date_month
            end_date_year = ed.year if ed else start_date_year + 1
        except (ValueError, TypeError):
            end_date_month = start_date_month
            end_date_year = start_date_year + 1

        return render_template(
            "timeline.html",
            proposal=proposal,
            tasks=proposal.tasks,
            start_date=proposal.start_date,
            start_date_month=start_date_month,
            start_date_year=start_date_year,
            end_date=proposal.end_date,
            end_date_month=end_date_month,
            end_date_year=end_date_year,
            task_budgets=task_budgets,
        )

    @app.route("/custom-sections/<proposal_id>")
    def custom_sections_tab(proposal_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404
        custom_sections = getattr(proposal, 'custom_sections', [])
        sections = sorted(custom_sections, key=lambda s: s.get("order", 0))
        return render_template(
            "custom_sections.html",
            proposal=proposal,
            sections=sections
        )

    @app.route("/api/section/<proposal_id>", methods=["POST"])
    def add_section(proposal_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404

        data = request.get_json()
        custom_sections = getattr(proposal, 'custom_sections', [])
        new_section = {
            "id": str(uuid.uuid4()),
            "title": data.get("title", "New Section"),
            "content": data.get("content", ""),
            "order": len(custom_sections)
        }
        if not hasattr(proposal, 'custom_sections') or proposal.custom_sections is None:
            proposal.custom_sections = []
        proposal.custom_sections.append(new_section)
        proposal.save()
        return jsonify(new_section), 201

    @app.route("/api/section/<proposal_id>/<section_id>", methods=["PUT"])
    def update_section(proposal_id, section_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404

        data = request.get_json()
        sections = getattr(proposal, 'custom_sections', [])
        for section in sections:
            if section["id"] == section_id:
                section.update({
                    "title": data.get("title", section["title"]),
                    "content": data.get("content", section["content"]),
                    "order": data.get("order", section.get("order", 0))
                })
                break

        proposal.save()
        return jsonify({"ok": True})

    @app.route("/api/section/<proposal_id>/<section_id>", methods=["DELETE"])
    def delete_section(proposal_id, section_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404

        sections = getattr(proposal, 'custom_sections', [])
        proposal.custom_sections = [
            s for s in sections if s["id"] != section_id
        ]
        proposal.save()
        return jsonify({"ok": True})

    @app.route("/api/section/<proposal_id>/import-excel", methods=["POST"])
    def import_excel_section(proposal_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404

        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({"error": "Only Excel files (.xlsx, .xls) are supported"}), 400

        try:
            import pandas as pd
            import io
        except ImportError:
            logger.error("Pandas/openpyxl not installed")
            return jsonify(ERROR_MESSAGES['EXCEL_NOT_INSTALLED']), 500

        try:
            # Read Excel file
            excel_data = file.read()
            excel_file = io.BytesIO(excel_data)
            df = pd.read_excel(excel_file, engine='openpyxl' if file.filename.endswith('.xlsx') else 'xlrd')

            # Convert DataFrame to markdown table
            markdown_content = df.to_markdown(index=False)

            # Create new section with Excel data
            custom_sections = getattr(proposal, 'custom_sections', [])
            new_section = {
                "id": str(uuid.uuid4()),
                "title": request.form.get('title', file.filename),
                "content": markdown_content,
                "order": len(custom_sections)
            }

            if not hasattr(proposal, 'custom_sections') or proposal.custom_sections is None:
                proposal.custom_sections = []
            proposal.custom_sections.append(new_section)
            proposal.save()
            
            logger.info(f"Imported Excel file as section: {new_section['id']}")
            return jsonify(new_section), 201

        except (ValueError, KeyError) as e:
            logger.warning(f"Invalid Excel file: {e}")
            return jsonify({**ERROR_MESSAGES['EXCEL_INVALID_FILE'], 'details': str(e)}), 400
        except Exception as e:
            logger.error(f"Excel import failed: {e}")
            return jsonify(ERROR_MESSAGES['EXCEL_PROCESSING_ERROR']), 500

    @app.route("/api/section/<proposal_id>/reorder", methods=["PUT"])
    def reorder_sections(proposal_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404

        data = request.get_json()
        section_order = data.get('section_order', [])

        sections = getattr(proposal, 'custom_sections', [])
        section_dict = {s['id']: s for s in sections}

        reordered = []
        for idx, section_id in enumerate(section_order):
            if section_id in section_dict:
                section = section_dict[section_id]
                section['order'] = idx
                reordered.append(section)

        proposal.custom_sections = reordered
        proposal.save()
        return jsonify({"ok": True})

    @app.route("/preview/<proposal_id>")
    def preview_tab(proposal_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404

        ctx = build_export_context(proposal)
        return render_template("preview.html", **ctx)

    @app.route("/api/task/<proposal_id>", methods=["POST"])
    def add_task(proposal_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404

        data = request.get_json()
        task = {
            "id": uuid.uuid4().hex[:8],
            "name": data.get("name", ""),
            "description": data.get("description", ""),
            "start_month": data.get("start_month", 1),
            "start_year": data.get("start_year", datetime.now().year),
            "duration_months": data.get("duration_months", 1),
            "lead_months": data.get("lead_months", 0),
            "lead_entity": data.get("lead_entity", ""),
            "recurring": data.get("recurring", False),
            "recurring_interval": data.get("recurring_interval", 3),
        }
        proposal.tasks.append(task)
        proposal.save()
        return jsonify(task), 201

    @app.route("/api/task/<proposal_id>/<task_id>", methods=["DELETE"])
    def delete_task(proposal_id, task_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404

        proposal.tasks = [t for t in proposal.tasks if t.get("id") != task_id]
        proposal.budget_items = [b for b in proposal.budget_items if b.get("task_id") != task_id]
        proposal.save()
        return "", 200

    @app.route("/api/budget/<proposal_id>", methods=["POST"])
    def add_budget_item(proposal_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404

        data = request.get_json()
        try:
            item = {
                "id": uuid.uuid4().hex[:8],
                "task_id": data.get("task_id", ""),
                "name": data.get("name", ""),
                "cost_per_unit": validate_numeric(data.get("cost_per_unit", 0), "cost_per_unit", min_val=0.0),
                "units": validate_numeric(data.get("units", 1), "units", min_val=0.0),
            }
        except ValueError as e:
            logger.warning(f"Invalid budget item data: {e}")
            return jsonify({**ERROR_MESSAGES['INVALID_NUMERIC'], 'details': str(e)}), 400
        
        proposal.budget_items.append(item)
        proposal.save()
        logger.debug(f"Added budget item {item['id']} to proposal {proposal_id}")
        return jsonify(item), 201

    @app.route("/api/budget/<proposal_id>/<item_id>", methods=["DELETE"])
    def delete_budget_item(proposal_id, item_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404

        proposal.budget_items = [b for b in proposal.budget_items if b.get("id") != item_id]
        proposal.save()
        return jsonify({"ok": True})

    @app.route("/api/budget/<proposal_id>/<item_id>", methods=["PUT"])
    def update_budget_item(proposal_id: str, item_id: str) -> Tuple[Response, int]:
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404

        data = request.get_json()
        if not data:
            return jsonify(ERROR_MESSAGES['NO_DATA']), 400

        for item in proposal.budget_items:
            if item.get("id") == item_id:
                try:
                    item["task_id"] = data.get("task_id", item.get("task_id", ""))
                    item["name"] = data.get("name", item.get("name", ""))
                    item["cost_per_unit"] = validate_numeric(
                        data.get("cost_per_unit", item.get("cost_per_unit", 0)),
                        "cost_per_unit",
                        min_val=0.0
                    )
                    item["units"] = validate_numeric(
                        data.get("units", item.get("units", 1)),
                        "units",
                        min_val=0.0
                    )
                except ValueError as e:
                    logger.warning(f"Invalid budget item data: {e}")
                    return jsonify({**ERROR_MESSAGES['INVALID_NUMERIC'], 'details': str(e)}), 400
                break
        else:
            return jsonify(ERROR_MESSAGES['BUDGET_ITEM_NOT_FOUND']), 404

        proposal.save()
        logger.debug(f"Updated budget item {item_id} in proposal {proposal_id}")
        return jsonify({"ok": True}), 200

    @app.route("/api/proposal/<proposal_id>/import-budget", methods=["POST"])
    def import_budget(proposal_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify(ERROR_MESSAGES['PROPOSAL_NOT_FOUND']), 404

        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({"error": "Only Excel files (.xlsx, .xls) are supported"}), 400

        try:
            from openpyxl import load_workbook
            import io
        except ImportError:
            return jsonify({"error": "openpyxl not installed"}), 500

        try:
            excel_data = file.read()
            wb = load_workbook(io.BytesIO(excel_data), read_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return jsonify({"error": "File has no data rows"}), 400

            header = [str(c).strip().lower() if c else "" for c in rows[0]]
            if not all(h in header for h in ["task", "item", "cost/unit", "units"]):
                return jsonify({"error": "Required columns: Task, Item, Cost/Unit, Units"}), 400

            task_idx = header.index("task")
            item_idx = header.index("item")
            cost_idx = header.index("cost/unit")
            units_idx = header.index("units")

            existing_tasks = {t["name"].strip().lower(): t for t in proposal.tasks}
            created_tasks = 0
            created_items = 0

            for row in rows[1:]:
                task_name = str(row[task_idx]).strip() if row[task_idx] else ""
                item_name = str(row[item_idx]).strip() if row[item_idx] else ""

                if not task_name or not item_name:
                    continue

                task_key = task_name.lower()
                if task_key not in existing_tasks:
                    new_task = {
                        "id": uuid.uuid4().hex[:8],
                        "name": task_name,
                        "description": "",
                        "start_month": 1,
                        "start_year": datetime.now().year,
                        "duration_months": 12,
                    }
                    proposal.tasks.append(new_task)
                    existing_tasks[task_key] = new_task
                    created_tasks += 1

                try:
                    cost = float(row[cost_idx]) if row[cost_idx] else 0
                    units = float(row[units_idx]) if row[units_idx] else 1
                except (ValueError, TypeError):
                    cost = 0
                    units = 1

                budget_item = {
                    "id": uuid.uuid4().hex[:8],
                    "task_id": existing_tasks[task_key]["id"],
                    "name": item_name,
                    "cost_per_unit": cost,
                    "units": units,
                }
                proposal.budget_items.append(budget_item)
                created_items += 1

            wb.close()
            proposal.save()

            return jsonify({
                "ok": True,
                "created_tasks": created_tasks,
                "created_items": created_items,
            }), 200

        except Exception as e:
            logger.error(f"Budget import failed: {e}")
            return jsonify({"error": f"Failed to process file: {str(e)}"}), 500

    @app.route("/api/budget-template", methods=["GET"])
    def download_budget_template():
        from openpyxl import Workbook
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = "Budget Template"
        ws.append(["Task", "Item", "Cost/Unit", "Units"])
        ws.append(["Scoping", "Initial meeting", 200, 6])
        ws.append(["Scoping", "Data collection", 200, 20])
        ws.append(["Analysis", "Forest valuation", 300, 40])
        ws.append(["Results", "Report writing", 250, 30])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=budget_template.xlsx"}
        )

    @app.route("/tracker/<proposal_id>")
    def tracker(proposal_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return redirect(url_for("index"))

        indirect_percent = getattr(proposal, 'indirect_percent', 0) or 0
        indirect_amount = proposal.total_budget * (indirect_percent / 100)
        total_with_indirect = proposal.total_budget + indirect_amount

        task_budgets = {}
        timings = proposal.budget_item_timings or {}
        for task in proposal.tasks:
            items = [b for b in proposal.budget_items if b.get("task_id") == task["id"]]
            for item in items:
                t = timings.get(item.get("id", ""), {})
                if t:
                    item["actual_cost"] = t.get("actual_cost", 0)
            subtotal = sum(i.get("cost_per_unit", 0) * i.get("units", 0) for i in items)
            actual_total = sum(i.get("actual_cost", 0) for i in items)
            task_budgets[task["id"]] = {
                "task": task,
                "items": items,
                "subtotal": subtotal,
                "actual_total": actual_total,
            }

        milestones = getattr(proposal, 'milestones', []) or []
        reports = getattr(proposal, 'reports', []) or []

        completed_tasks = sum(1 for t in proposal.tasks if t.get("status") == "completed")
        total_tasks = len(proposal.tasks)
        overall_pct = round(completed_tasks / total_tasks * 100) if total_tasks else 0

        return render_template(
            "tracker.html",
            proposal=proposal,
            tasks=proposal.tasks,
            task_budgets=task_budgets,
            total_budget=proposal.total_budget,
            indirect_percent=indirect_percent,
            indirect_amount=indirect_amount,
            total_with_indirect=total_with_indirect,
            milestones=milestones,
            reports=reports,
            overall_pct=overall_pct,
        )

    @app.route("/api/tracker/<proposal_id>/task/<task_id>", methods=["PUT"])
    def update_tracker_task(proposal_id, task_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify({"error": "Not found"}), 404

        data = request.get_json()
        for task in proposal.tasks:
            if task.get("id") == task_id:
                if "status" in data:
                    task["status"] = data["status"]
                if "progress_pct" in data:
                    task["progress_pct"] = int(data["progress_pct"])
                if "actual_start" in data:
                    task["actual_start"] = data["actual_start"]
                if "actual_end" in data:
                    task["actual_end"] = data["actual_end"]
                if "notes" in data:
                    task["notes"] = data["notes"]
                break
        else:
            return jsonify({"error": "Task not found"}), 404

        proposal.save()
        return jsonify({"ok": True})

    @app.route("/api/tracker/<proposal_id>/budget/<item_id>", methods=["PUT"])
    def update_tracker_budget(proposal_id, item_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify({"error": "Not found"}), 404

        data = request.get_json()
        timings = proposal.budget_item_timings or {}
        if item_id not in timings:
            timings[item_id] = {}
        if "actual_cost" in data:
            timings[item_id]["actual_cost"] = float(data["actual_cost"])
        proposal.budget_item_timings = timings
        proposal.save()
        return jsonify({"ok": True})

    @app.route("/api/tracker/<proposal_id>/milestone", methods=["POST"])
    def add_milestone(proposal_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify({"error": "Not found"}), 404

        data = request.get_json()
        milestones = getattr(proposal, 'milestones', []) or []
        milestone = {
            "id": uuid.uuid4().hex[:8],
            "name": data.get("name", ""),
            "date": data.get("date", ""),
            "completed": False,
        }
        milestones.append(milestone)
        proposal.milestones = milestones
        proposal.save()
        return jsonify(milestone), 201

    @app.route("/api/tracker/<proposal_id>/milestone/<milestone_id>", methods=["PUT"])
    def update_milestone(proposal_id, milestone_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify({"error": "Not found"}), 404

        data = request.get_json()
        milestones = getattr(proposal, 'milestones', []) or []
        for m in milestones:
            if m["id"] == milestone_id:
                if "name" in data:
                    m["name"] = data["name"]
                if "date" in data:
                    m["date"] = data["date"]
                if "completed" in data:
                    m["completed"] = data["completed"]
                break
        else:
            return jsonify({"error": "Milestone not found"}), 404

        proposal.milestones = milestones
        proposal.save()
        return jsonify({"ok": True})

    @app.route("/api/tracker/<proposal_id>/milestone/<milestone_id>", methods=["DELETE"])
    def delete_milestone(proposal_id, milestone_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify({"error": "Not found"}), 404

        milestones = getattr(proposal, 'milestones', []) or []
        proposal.milestones = [m for m in milestones if m["id"] != milestone_id]
        proposal.save()
        return jsonify({"ok": True})

    @app.route("/api/tracker/<proposal_id>/report", methods=["POST"])
    def add_report(proposal_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify({"error": "Not found"}), 404

        data = request.get_json()
        reports = getattr(proposal, 'reports', []) or []
        report = {
            "id": uuid.uuid4().hex[:8],
            "date": data.get("date", datetime.now().strftime("%Y-%m-%d")),
            "title": data.get("title", ""),
            "content": data.get("content", ""),
        }
        reports.append(report)
        proposal.reports = reports
        proposal.save()
        return jsonify(report), 201

    @app.route("/api/tracker/<proposal_id>/report/<report_id>", methods=["PUT"])
    def update_report(proposal_id, report_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify({"error": "Not found"}), 404

        data = request.get_json()
        reports = getattr(proposal, 'reports', []) or []
        for r in reports:
            if r["id"] == report_id:
                if "title" in data:
                    r["title"] = data["title"]
                if "date" in data:
                    r["date"] = data["date"]
                if "content" in data:
                    r["content"] = data["content"]
                break
        else:
            return jsonify({"error": "Report not found"}), 404

        proposal.reports = reports
        proposal.save()
        return jsonify({"ok": True})

    @app.route("/api/tracker/<proposal_id>/report/<report_id>", methods=["DELETE"])
    def delete_report(proposal_id, report_id):
        proposal = Proposal.load(proposal_id)
        if not proposal:
            return jsonify({"error": "Not found"}), 404

        reports = getattr(proposal, 'reports', []) or []
        proposal.reports = [r for r in reports if r["id"] != report_id]
        proposal.save()
        return jsonify({"ok": True})

    return app


def run_server() -> None:
    """Run the development server."""
    import logging as _logging
    _logging.getLogger("werkzeug").setLevel(_logging.WARNING)
    app = create_app()
    app.run(debug=Config.DEBUG, host=Config.HOST, port=Config.PORT)


if __name__ == "__main__":
    run_server()
